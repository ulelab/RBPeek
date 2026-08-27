#!/usr/bin/env python3
"""
Heatmap of normalised expression per gene across samples, for a gene list.

Takes a counts/expression matrix (genes x samples) and a gene list, and draws a clustered
heatmap. If the matrix holds raw counts, supply a GTF with --gtf and the script computes
RPKM; if it already holds RPKM/TPM/normalised counts, pass --already-normalised.

  python3 scripts/plot_expression_heatmap.py \\
    -c counts.tsv --gtf /camp/.../filtered.gencode.v39.main.annotation.gtf \\
    -g DESeq2_...xlsx --gene-column 1 \\
    -o results/centro/expression_heatmap.png

Gene identifiers are matched case-insensitively after stripping whitespace. Anything in the
gene list that is absent from the matrix is reported rather than silently dropped - including
Excel's date-mangled symbols (MARCH1 -> "2025-03-01"), which never match anything.
"""

import argparse
import csv
import gzip
import os
import re
import sys

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import seaborn as sns


def parse_args():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("-c", "--counts", required=True,
                   help="Matrix of genes x samples (TSV/CSV). First column = gene id/symbol.")
    p.add_argument("-g", "--genes", required=True,
                   help="Gene list: .xlsx (first column used), or a plain text file, one per line.")
    p.add_argument("--gene-column", type=int, default=1,
                   help="1-based column of the gene list file to read (default 1)")
    p.add_argument("--gtf", default=None,
                   help="GTF for gene lengths, to convert raw counts to RPKM. Omit with --already-normalised.")
    p.add_argument("--already-normalised", action="store_true",
                   help="Matrix is already RPKM/TPM/normalised; skip length normalisation.")
    p.add_argument("--samples", default=None,
                   help="Comma-separated sample columns to keep, in this order (default: all numeric columns)")
    p.add_argument("--top-n", type=int, default=50,
                   help="If more genes match than this, keep the most variable N (default 50; 0 = keep all)")
    p.add_argument("--scale", choices=["zscore", "log2", "none"], default="zscore",
                   help="Row scaling. zscore (default) compares SHAPE across samples; log2 keeps magnitude.")
    p.add_argument("-o", "--out", required=True, help="Output PNG")
    p.add_argument("--out-table", default=None, help="Also write the plotted matrix as TSV")
    return p.parse_args()


def _open(path):
    return gzip.open(path, "rt") if str(path).endswith(".gz") else open(path)


def read_gene_list(path, col):
    if str(path).lower().endswith((".xlsx", ".xlsm")):
        try:
            import openpyxl
        except ImportError:
            sys.exit("openpyxl needed to read %s - use a plain text gene list instead" % path)
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        out = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # header
            if len(row) >= col and row[col - 1] is not None:
                out.append(str(row[col - 1]).strip())
        return out
    with _open(path) as fh:
        return [l.split("\t")[col - 1].strip() for l in fh if l.strip()]


def read_matrix(path, samples):
    delim = "," if str(path).lower().endswith(".csv") else "\t"
    with _open(path) as fh:
        reader = csv.reader(fh, delimiter=delim)
        header = next(reader)
        rows, genes = [], []
        for r in reader:
            if not r or not r[0].strip():
                continue
            genes.append(r[0].strip())
            rows.append(r[1:])
    cols = header[1:]
    keep = list(range(len(cols)))
    if samples:
        want = [s.strip() for s in samples.split(",")]
        missing = [w for w in want if w not in cols]
        if missing:
            sys.exit("sample column(s) not in matrix: %s\navailable: %s" % (", ".join(missing), ", ".join(cols)))
        keep = [cols.index(w) for w in want]
    M = np.full((len(rows), len(keep)), np.nan)
    for i, r in enumerate(rows):
        for j, k in enumerate(keep):
            try:
                M[i, j] = float(r[k])
            except (ValueError, IndexError):
                M[i, j] = np.nan
    return genes, [cols[k] for k in keep], M


def gene_lengths_from_gtf(path):
    """Union of exons per gene symbol, in bp - the standard denominator for RPKM."""
    ivs = {}
    with _open(path) as fh:
        for line in fh:
            if line.startswith("#"):
                continue
            f = line.rstrip("\n").split("\t")
            if len(f) < 9 or f[2] != "exon":
                continue
            m = re.search(r'gene_name "([^"]+)"', f[8])
            if not m:
                continue
            ivs.setdefault(m.group(1), []).append((int(f[3]) - 1, int(f[4])))
    lengths = {}
    for g, spans in ivs.items():
        spans.sort()
        total, cur_s, cur_e = 0, None, None
        for s, e in spans:
            if cur_e is None or s > cur_e:
                if cur_e is not None:
                    total += cur_e - cur_s
                cur_s, cur_e = s, e
            else:
                cur_e = max(cur_e, e)
        if cur_e is not None:
            total += cur_e - cur_s
        lengths[g] = total
    return lengths


def main():
    args = parse_args()
    if not args.already_normalised and not args.gtf:
        sys.exit("supply --gtf to convert raw counts to RPKM, or pass --already-normalised")

    wanted = read_gene_list(args.genes, args.gene_column)
    genes, samples, M = read_matrix(args.counts, args.samples)
    print("matrix: %d genes x %d samples (%s)" % (len(genes), len(samples), ", ".join(samples)))
    print("gene list: %d entries" % len(wanted))

    index = {}
    for i, g in enumerate(genes):
        index.setdefault(g.strip().upper(), i)
    hits, missing = [], []
    seen = set()
    for g in wanted:
        k = g.upper()
        if k in index and k not in seen:
            seen.add(k)
            hits.append((g, index[k]))
        elif k not in index:
            missing.append(g)
    print("matched %d of %d gene-list entries" % (len(hits), len(wanted)))
    if missing:
        dates = [m for m in missing if re.match(r"^\d{4}-\d{2}-\d{2}", m)]
        print("  unmatched: %d%s" % (len(missing), "" if not dates else
              "  (%d are Excel date-mangled symbols, e.g. %s - these are MARCH/SEPT genes and "
              "will never match)" % (len(dates), dates[0])))
        print("  first few: %s" % ", ".join(missing[:8]))
    if not hits:
        sys.exit("no genes matched - check that the matrix and the gene list use the same identifier type")

    names = [g for g, _ in hits]
    X = M[[i for _, i in hits], :]
    ok = ~np.all(np.isnan(X), axis=1)
    names = [n for n, k in zip(names, ok) if k]
    X = X[ok]
    X = np.nan_to_num(X, nan=0.0)

    if not args.already_normalised:
        lengths = gene_lengths_from_gtf(args.gtf)
        lg = np.array([lengths.get(n.upper(), lengths.get(n, 0)) for n in names], dtype=float)
        # try case-insensitively for anything still zero
        if (lg == 0).any():
            up = dict((k.upper(), v) for k, v in lengths.items())
            for i, n in enumerate(names):
                if lg[i] == 0:
                    lg[i] = up.get(n.upper(), 0)
        nolen = int((lg == 0).sum())
        if nolen:
            print("  %d gene(s) had no exons in the GTF and are dropped from RPKM" % nolen)
        keep = lg > 0
        names = [n for n, k in zip(names, keep) if k]
        X, lg = X[keep], lg[keep]
        libs = X.sum(axis=0)
        libs[libs == 0] = 1.0
        X = (X * 1e9) / (lg[:, None] * libs[None, :])
        print("  converted raw counts to RPKM (library sizes: %s)"
              % ", ".join("%.1fM" % (l / 1e6) for l in libs))

    if args.top_n and X.shape[0] > args.top_n:
        v = np.log1p(X).std(axis=1)
        idx = np.argsort(-v)[: args.top_n]
        idx = np.sort(idx)
        print("  %d genes matched; plotting the %d most variable (use --top-n 0 for all)"
              % (X.shape[0], args.top_n))
        names = [names[i] for i in idx]
        X = X[idx]

    D = np.log2(X + 1)
    label = "log2(RPKM+1)" if not args.already_normalised else "log2(value+1)"
    if args.scale == "zscore":
        sd = D.std(axis=1, keepdims=True)
        sd[sd == 0] = 1.0
        D = (D - D.mean(axis=1, keepdims=True)) / sd
        label = "row z-score\n(%s)" % label
    elif args.scale == "none":
        D = X
        label = "RPKM" if not args.already_normalised else "value"

    h = max(4.0, 0.18 * len(names))
    cg = sns.clustermap(D, cmap="vlag" if args.scale == "zscore" else "magma",
                        col_cluster=len(samples) > 2, row_cluster=len(names) > 1,
                        xticklabels=samples, yticklabels=names,
                        figsize=(max(6, 1.1 * len(samples) + 4), h),
                        cbar_kws={"label": label})
    cg.ax_heatmap.set_xlabel("Sample")
    cg.ax_heatmap.set_ylabel("Gene")
    cg.ax_heatmap.tick_params(axis="y", labelsize=max(4, min(9, 400.0 / max(len(names), 1))))
    plt.setp(cg.ax_heatmap.get_xticklabels(), rotation=90)
    d = os.path.dirname(args.out)
    if d:
        os.makedirs(d, exist_ok=True)
    cg.savefig(args.out, dpi=200, bbox_inches="tight")
    print("wrote %s (%d genes x %d samples)" % (args.out, len(names), len(samples)))

    if args.out_table:
        with open(args.out_table, "w") as fh:
            fh.write("gene\t" + "\t".join(samples) + "\n")
            for n, row in zip(names, X):
                fh.write(n + "\t" + "\t".join("%.6g" % v for v in row) + "\n")
        print("wrote %s" % args.out_table)


if __name__ == "__main__":
    main()
